import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# 读取原始Excel
xl = pd.ExcelFile('C:/Users/ADMIN/Desktop/Ratecard2024.xlsx')

# 创建新的工作簿
wb = Workbook()
wb.remove(wb.active)

# 定义样式
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
subheader_fill = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")
subheader_font = Font(bold=True, size=10)
data_font = Font(size=10)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

def apply_styles(ws, max_row, max_col):
    col_widths = [25] + [15] * (max_col - 1)
    for i, width in enumerate(col_widths[:max_col], 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    for row in range(1, max_row + 1):
        ws.row_dimensions[row].height = 20
    
    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.border = border
            if cell.row == 1:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
            elif cell.row == 2:
                cell.fill = subheader_fill
                cell.font = subheader_font
                cell.alignment = center_align
            else:
                cell.font = data_font
                cell.alignment = center_align

def process_ipsos(df, sheet_name):
    """处理IPSOS sheet"""
    rows_data = []
    loi_values = ['<=5', '<=10', '<=15', '<=20', '<=25', '<=30', '<=35', '<=40', '<=45']
    
    # 从第4行开始是数据(索引3)，IR在第一列
    for i in range(3, 14):
        ir_val = df.iloc[i, 0]
        if pd.notna(ir_val):
            row_data = {'IR(渗透率)': str(ir_val)}
            for j, loi in enumerate(loi_values):
                val = df.iloc[i, j+1]
                if pd.notna(val):
                    try:
                        row_data[f'LOI {loi}'] = round(float(val), 2)
                    except:
                        row_data[f'LOI {loi}'] = str(val)
                else:
                    row_data[f'LOI {loi}'] = ''
            rows_data.append(row_data)
    
    return rows_data, loi_values

def process_nielsen(df, sheet_name):
    """处理Nielsen sheet"""
    rows_data = []
    # LOI在第2行（索引2）
    loi_row = df.iloc[2]
    loi_values = []
    for val in loi_row[2:]:
        if pd.notna(val):
            loi_values.append(str(val).strip())
    
    # 数据从第3行开始
    for i in range(3, len(df)):
        ir_val = df.iloc[i, 1]  # IR在第1列
        if pd.notna(ir_val):
            row_data = {'IR(渗透率)': str(ir_val).strip()}
            for j, loi in enumerate(loi_values):
                val = df.iloc[i, j+2]
                if pd.notna(val):
                    try:
                        row_data[f'LOI {loi}'] = round(float(val), 2)
                    except:
                        row_data[f'LOI {loi}'] = str(val)
                else:
                    row_data[f'LOI {loi}'] = ''
            rows_data.append(row_data)
    
    return rows_data, loi_values

def process_rakuten(df, sheet_name):
    """处理Rakuten sheet"""
    rows_data = []
    # LOI在第2行（索引2）
    loi_row = df.iloc[2]
    loi_values = []
    for val in loi_row[2:]:
        if pd.notna(val):
            loi_values.append(str(val).strip())
    
    # 数据从第3行开始
    for i in range(3, len(df)):
        ir_val = df.iloc[i, 1]  # IR在第1列
        if pd.notna(ir_val):
            row_data = {'IR(渗透率)': str(ir_val).strip()}
            for j, loi in enumerate(loi_values):
                val = df.iloc[i, j+2]
                if pd.notna(val):
                    try:
                        row_data[f'LOI {loi}'] = round(float(val), 2)
                    except:
                        row_data[f'LOI {loi}'] = str(val)
                else:
                    row_data[f'LOI {loi}'] = ''
            rows_data.append(row_data)
    
    return rows_data, loi_values

def process_macromill(df, sheet_name):
    """处理Macromill JP sheet"""
    rows_data = []
    # LOI在第2行（索引2）
    loi_row = df.iloc[2]
    loi_values = []
    for val in loi_row[2:]:
        if pd.notna(val):
            loi_values.append(str(val).strip())
    
    # 数据从第3行开始
    for i in range(3, len(df)):
        ir_val = df.iloc[i, 1]  # IR在第1列
        if pd.notna(ir_val):
            row_data = {'IR(渗透率)': str(ir_val).strip()}
            for j, loi in enumerate(loi_values):
                val = df.iloc[i, j+2]
                if pd.notna(val):
                    try:
                        row_data[f'LOI {loi}'] = round(float(val), 2)
                    except:
                        row_data[f'LOI {loi}'] = str(val)
                else:
                    row_data[f'LOI {loi}'] = ''
            rows_data.append(row_data)
    
    return rows_data, loi_values

def process_kantar_cn(df, sheet_name):
    """处理Kantar CN sheet"""
    rows_data = []
    # LOI在第1行（索引1）
    loi_row = df.iloc[1]
    loi_values = []
    for val in loi_row[1:]:
        if pd.notna(val):
            loi_values.append(str(val).strip())
    
    # 数据从第2行开始
    for i in range(2, len(df)):
        ir_val = df.iloc[i, 0]  # IR在第0列
        if pd.notna(ir_val):
            row_data = {'IR(渗透率)': str(ir_val).strip()}
            for j, loi in enumerate(loi_values):
                val = df.iloc[i, j+1]
                if pd.notna(val):
                    try:
                        row_data[f'LOI {loi}'] = round(float(val), 2)
                    except:
                        row_data[f'LOI {loi}'] = str(val)
                else:
                    row_data[f'LOI {loi}'] = ''
            rows_data.append(row_data)
    
    return rows_data, loi_values

def process_glg(df, sheet_name):
    """处理GLG sheet - 特殊格式"""
    rows_data = []
    
    # GLG的格式：AGE在第0列，数据从第4行开始
    # LOI在第11列（最后一列）
    # 数据列：1-10列是不同IR的值
    
    loi_values = ['1~5', '5~10', '10~15', '15~20', '20~30']
    
    for i in range(4, len(df)):
        age_val = df.iloc[i, 0]
        if pd.notna(age_val):
            age_str = str(age_val).strip()
            # 检查是否是需要跳过的行
            if 'Parent Referral' in age_str:
                continue
                
            row_data = {'AGE': age_str}
            
            # 只取前5个数据列的值
            for j in range(5):
                val = df.iloc[i, j+1]
                if pd.notna(val):
                    try:
                        row_data[f'LOI {loi_values[j]}'] = round(float(val), 2)
                    except:
                        row_data[f'LOI {loi_values[j]}'] = str(val)
                else:
                    row_data[f'LOI {loi_values[j]}'] = ''
            
            rows_data.append(row_data)
    
    return rows_data, loi_values

def process_ipsos_b2b(df, sheet_name):
    """处理IPSOS B2B sheet"""
    rows_data = []
    # LOI在第2行（索引2）
    loi_row = df.iloc[2]
    loi_values = []
    for val in loi_row[1:]:
        if pd.notna(val):
            loi_values.append(str(val).strip())
    
    # 数据从第3行开始，受访者类型在第0列
    for i in range(3, len(df)):
        type_val = df.iloc[i, 0]
        if pd.notna(type_val):
            row_data = {'受访者类型': str(type_val).strip()}
            for j, loi in enumerate(loi_values):
                val = df.iloc[i, j+1]
                if pd.notna(val):
                    try:
                        row_data[f'LOI {loi}'] = round(float(val), 2)
                    except:
                        row_data[f'LOI {loi}'] = str(val)
                else:
                    row_data[f'LOI {loi}'] = ''
            rows_data.append(row_data)
    
    return rows_data, loi_values

# 创建汇总sheet
ws_summary = wb.create_sheet("汇总目录")
ws_summary.append(["Ratecard 2024 汇总目录"])
ws_summary.merge_cells('A1:D1')
ws_summary['A1'].font = Font(bold=True, size=14)
ws_summary['A1'].alignment = center_align
ws_summary.row_dimensions[1].height = 25
ws_summary.append([""])
ws_summary.append(["数据来源Sheet列表"])
ws_summary.append(["序号", "Sheet名称", "数据类型", "说明"])
ws_summary.append([1, "IPSOS", "定价矩阵", "不含税报价"])
ws_summary.append([2, "Nielsen", "定价矩阵", "含税报价"])
ws_summary.append([3, "Rakuten", "定价矩阵", "样本定价"])
ws_summary.append([4, "Macromill JP", "定价矩阵", "日本样本"])
ws_summary.append([5, "Kantar CN", "定价矩阵", "中国样本"])
ws_summary.append([6, "GLG", "定价矩阵", "专家访谈"])
ws_summary.append([7, "IPSOS B2B", "B2B定价", "B2B样本"])
ws_summary.append([8, "Quotation", "报价单", "项目报价"])
apply_styles(ws_summary, 12, 4)

# 定义处理器
processors = {
    'IPSOS': process_ipsos,
    'Nielsen': process_nielsen,
    'Rakuten': process_rakuten,
    'Macromill JP': process_macromill,
    'Kantar CN': process_kantar_cn,
    'GLG': process_glg,
    'IPSOS B2B': process_ipsos_b2b,
}

# 处理每个sheet
for sheet_name in processors.keys():
    df = pd.read_excel(xl, sheet_name=sheet_name, header=None)
    
    if df.empty:
        continue
    
    processor = processors[sheet_name]
    rows_data, loi_values = processor(df, sheet_name)
    
    ws = wb.create_sheet(sheet_name)
    
    if sheet_name == 'IPSOS B2B':
        headers = ['受访者类型'] + [f'LOI {loi}' for loi in loi_values]
    elif sheet_name == 'GLG':
        headers = ['AGE'] + [f'LOI {loi}' for loi in loi_values]
    else:
        headers = ['IR(渗透率)'] + [f'LOI {loi}' for loi in loi_values]
    
    ws.append(['数据来源: ' + sheet_name] + [''] * (len(headers) - 1))
    ws.append(headers)
    
    for row_data in rows_data:
        if sheet_name == 'IPSOS B2B':
            row = [row_data.get('受访者类型', '')]
        elif sheet_name == 'GLG':
            row = [row_data.get('AGE', '')]
        else:
            row = [row_data.get('IR(渗透率)', '')]
        for loi in loi_values:
            row.append(row_data.get(f'LOI {loi}', ''))
        ws.append(row)
    
    apply_styles(ws, ws.max_row, len(headers))

# 处理Quotation
df_quote = pd.read_excel(xl, sheet_name='Quotation')
ws = wb.create_sheet('Quotation')
headers = ['Option', 'LOI(min)', 'IR(%)', 'N Required', 'N Feasible', 'CPI', 'Amount', 'FW days']
ws.append(['数据来源: Quotation'] + [''] * (len(headers) - 1))
ws.append(headers)

for idx in range(1, len(df_quote)):
    row = df_quote.iloc[idx]
    if pd.isna(row.iloc[0]) or str(row.iloc[0]) == 'Option':
        continue
    row_data = []
    for i in range(len(headers)):
        val = row.iloc[i] if i < len(row) else None
        row_data.append(val if pd.notna(val) else '')
    ws.append(row_data)

apply_styles(ws, ws.max_row, len(headers))

# 保存文件
output_path = 'C:/Users/ADMIN/Desktop/Ratecard2024_整理.xlsx'
wb.save(output_path)
print(f"文件已保存到: {output_path}")
print(f"Sheet列表: {wb.sheetnames}")
